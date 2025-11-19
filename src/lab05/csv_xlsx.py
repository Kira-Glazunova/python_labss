from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)

    if not csv_path.exists():
        raise FileNotFoundError("CSV файл не найден")
    
    # Создаем директорию для выходного файла
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    
    if csv_path.suffix != '.csv':
        raise ValueError("Неверный тип файла")

    with open(csv_path, 'r', encoding='utf-8') as csv_file:
        reader = csv.DictReader(csv_file)
        csv_data = list(reader)
        
        if len(csv_data) == 0:
            raise ValueError("Пустой файл")
        
        if not reader.fieldnames:
            raise ValueError("Файл не содержит заголовков")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.append(reader.fieldnames)
        for row in csv_data:
            ws.append([row[field] for field in reader.fieldnames])
        
        for column in ws.columns:
            max_length = 8
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(len(str(cell.value)), max_length)
            ws.column_dimensions[column_letter].width = max_length
        
        wb.save(xlsx_path)

# Пример вызова функции
if __name__ == "__main__":
    csv_to_xlsx('data/samples/people.csv', 'data/out/people_from_csv.xlsx')